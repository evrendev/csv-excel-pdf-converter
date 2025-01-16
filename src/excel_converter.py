import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

GERMAN_HEADERS = {
    'amount': 'Betrag',
    'project': 'Projekt',
    'first_name': 'Vorname',
    'last_name': 'Nachname',
    'Öffentlich Spender-Name (als Nachweis auf den Spendenbannern)': 'Öffentlicher Spendername'
}

def convert_to_excel(df, output_path):
    # Rename columns to German
    df_german = df.rename(columns=GERMAN_HEADERS)
    
    # Calculate total
    total = df['amount'].astype(float).sum()
    
    # Add total row
    df_german.loc[len(df_german)] = ['Gesamt', total, '', '', '']
    
    # Save to Excel
    df_german.to_excel(output_path, index=False)
    
    # Load workbook for formatting
    wb = load_workbook(output_path)
    ws = wb.active
    
    # Format headers (white text)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Format total row
    last_row = ws.max_row
    total_font = Font(bold=True)
    total_border = Border(top=Side(style='medium'))
    
    for cell in ws[last_row]:
        cell.font = total_font
        cell.border = total_border
    
    # Format amount column and total as currency
    for row in ws.iter_rows(min_row=2):
        if row[0].value and isinstance(row[0].value, (int, float)):
            row[0].number_format = '#,##0.00 €'
            row[0].alignment = Alignment(horizontal='right')
    
    # Create table (excluding total row)
    tab = Table(displayName="DonationsTable", ref=f"A1:E{last_row-1}")
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)
    
    # Save formatted workbook
    wb.save(output_path)
    
    return total